from __future__ import annotations

import csv
from pathlib import Path


def extract_text_for_path(path: Path) -> str | None:
    """Return extracted text for a file, or None if unsupported/unreadable.

    This is intentionally best-effort and dependency-light:
    - Text formats are read as UTF-8 (errors ignored)
    - PDF extraction uses pypdf if installed
    - XLSX extraction uses openpyxl if installed
    """
    ext = path.suffix.lower()
    if ext in {".txt", ".md", ".rst", ".json", ".yaml", ".yml", ".toml", ".sql", ".py", ".js", ".jsx", ".ts", ".tsx"}:
        return _read_text(path)
    if ext in {".csv", ".tsv"}:
        return _read_delimited(path, delimiter="," if ext == ".csv" else "\t")
    if ext == ".pdf":
        return _read_pdf(path)
    if ext == ".xlsx":
        return _read_xlsx(path)
    return None


def _read_text(path: Path) -> str | None:
    try:
        return path.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return None


def _read_delimited(path: Path, *, delimiter: str) -> str | None:
    try:
        raw = path.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return None

    # Normalize into a simple “table-ish” textual representation.
    out_lines: list[str] = []
    try:
        reader = csv.reader(raw.splitlines(), delimiter=delimiter)
        for row in reader:
            if not row:
                continue
            out_lines.append("\t".join(str(c).strip() for c in row))
    except Exception:
        return raw
    return "\n".join(out_lines)


def _read_pdf(path: Path) -> str | None:
    try:
        from pypdf import PdfReader
    except Exception:
        return None

    try:
        reader = PdfReader(str(path))
    except Exception:
        return None

    parts: list[str] = []
    for i, page in enumerate(getattr(reader, "pages", []) or []):
        try:
            txt = page.extract_text() or ""
        except Exception:
            txt = ""
        if not txt.strip():
            continue
        parts.append(f"\n\n--- page {i + 1} ---\n\n{txt.strip()}\n")
    joined = "\n".join(parts).strip()
    return joined or ""


def _read_xlsx(path: Path) -> str | None:
    try:
        from openpyxl import load_workbook
    except Exception:
        return None

    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    except Exception:
        return None

    out_lines: list[str] = []
    try:
        for ws in wb.worksheets:
            title = str(getattr(ws, "title", "") or "").strip() or "Sheet"
            out_lines.append(f"\n\n--- sheet {title} ---\n")
            try:
                for row in ws.iter_rows(values_only=True):
                    if not row:
                        continue
                    cells = [("" if c is None else str(c)).strip() for c in row]
                    if not any(cells):
                        continue
                    out_lines.append("\t".join(cells))
            except Exception:
                continue
    finally:
        try:
            wb.close()
        except Exception:
            pass

    joined = "\n".join(out_lines).strip()
    return joined or ""

